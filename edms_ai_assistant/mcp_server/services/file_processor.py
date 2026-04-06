# mcp-server/services/file_processor.py
"""
Сервис извлечения текста из файлов.
Перенесён из edms_ai_assistant/services/file_processor.py.

Поддерживаемые форматы: PDF, DOCX, DOC, TXT, XLSX, XLS.
"""
from __future__ import annotations

import asyncio
import logging
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_executor = ThreadPoolExecutor(max_workers=4)


def _extract_doc_via_fitz(file_path: str) -> str:
    """Извлечение текста из .doc/.docx через PyMuPDF."""
    import fitz  # type: ignore[import]

    doc = fitz.open(file_path)
    pages_text: list[str] = []
    for page in doc:
        text = page.get_text("text")
        if text and text.strip():
            pages_text.append(text.strip())
    doc.close()
    return "\n\n".join(pages_text)


def _extract_doc_via_mammoth(file_path: str) -> str:
    """Извлечение текста из .doc/.docx через mammoth."""
    import mammoth  # type: ignore[import]

    with open(file_path, "rb") as f:
        result = mammoth.extract_raw_text(f)
    return result.value.strip()


def _extract_docx_via_docx2txt(file_path: str) -> str:
    """Извлечение текста из .docx через docx2txt."""
    import docx2txt  # type: ignore[import]

    return docx2txt.process(file_path) or ""


class FileProcessorService:
    """
    Сервис извлечения текста из файлов различных форматов.

    Стратегии по расширению:
        .pdf  → PyPDFLoader
        .docx → docx2txt → mammoth → fitz (цепочка fallback)
        .doc  → fitz → mammoth (цепочка fallback)
        .txt  → TextLoader
        .xlsx → openpyxl
        .xls  → xlrd
    """

    SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".txt", ".xlsx", ".xls"}

    @classmethod
    async def extract_text_async(cls, file_path: str) -> str:
        """Async извлечение текста — CPU-работа делегируется в thread pool."""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(_executor, cls.extract_text, file_path)

    @classmethod
    def extract_text(cls, file_path: str) -> str:
        """Синхронное извлечение текста с выбором стратегии по расширению."""
        path = Path(file_path)

        if not path.exists():
            return f"Ошибка: файл не найден: {file_path}"

        ext = path.suffix.lower()
        if ext not in cls.SUPPORTED_EXTENSIONS:
            return f"Формат файла {ext} не поддерживается для анализа."

        try:
            if ext in (".xlsx", ".xls"):
                return cls._extract_from_excel(file_path, ext)
            if ext == ".doc":
                return cls._extract_doc(file_path)
            if ext == ".docx":
                return cls._extract_docx(file_path)
            if ext == ".pdf":
                return cls._extract_pdf(file_path)
            if ext == ".txt":
                return cls._extract_txt(file_path)
            return f"Формат {ext} не поддерживается."
        except Exception as e:
            logger.error("Ошибка чтения файла %s: %s", file_path, e, exc_info=True)
            return f"Ошибка: техническая ошибка при чтении файла {ext}: {e!s}"

    @classmethod
    def _extract_doc(cls, file_path: str) -> str:
        """Извлечение .doc через fitz → mammoth."""
        try:
            text = _extract_doc_via_fitz(file_path)
            if text and len(text.strip()) > 10:
                return text
        except Exception as e:
            logger.warning("fitz failed for .doc '%s': %s", file_path, e)

        try:
            text = _extract_doc_via_mammoth(file_path)
            if text and len(text.strip()) > 10:
                return text
        except Exception as e:
            logger.warning("mammoth failed for .doc '%s': %s", file_path, e)

        return (
            "Не удалось извлечь текст из .doc. "
            "Попробуйте пересохранить в формате .docx."
        )

    @classmethod
    def _extract_docx(cls, file_path: str) -> str:
        """Извлечение .docx через docx2txt → mammoth → fitz."""
        try:
            text = _extract_docx_via_docx2txt(file_path)
            if text and len(text.strip()) > 10:
                return text
        except Exception as e:
            logger.warning("docx2txt failed for '%s': %s", file_path, e)

        try:
            text = _extract_doc_via_mammoth(file_path)
            if text and len(text.strip()) > 10:
                return text
        except Exception as e:
            logger.warning("mammoth failed for '%s': %s", file_path, e)

        try:
            text = _extract_doc_via_fitz(file_path)
            if text and len(text.strip()) > 10:
                return text
        except Exception as e:
            logger.error("fitz failed for '%s': %s", file_path, e)

        return "Не удалось извлечь текст из .docx."

    @classmethod
    def _extract_pdf(cls, file_path: str) -> str:
        """Извлечение .pdf через LangChain PyPDFLoader."""
        from langchain_community.document_loaders import PyPDFLoader

        loader = PyPDFLoader(file_path)
        docs = loader.load()
        if not docs:
            return "Файл прочитан, но текстового содержимого не обнаружено."
        text = "\n\n".join(doc.page_content for doc in docs).strip()
        if not text:
            return "PDF состоит из изображений без текстового слоя."
        return text

    @classmethod
    def _extract_txt(cls, file_path: str) -> str:
        """Извлечение .txt через LangChain TextLoader."""
        from langchain_community.document_loaders import TextLoader

        loader = TextLoader(file_path, encoding="utf-8")
        docs = loader.load()
        if not docs:
            return "Текстовый файл пуст."
        return "\n\n".join(doc.page_content for doc in docs).strip()

    @classmethod
    def _extract_from_excel(cls, file_path: str, ext: str) -> str:
        """Извлечение таблиц из Excel в текстовый формат."""
        try:
            if ext == ".xlsx":
                import openpyxl

                wb = openpyxl.load_workbook(file_path, data_only=True)
                extracted_text = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    extracted_text.append(f"\n{'='*50}\nЛИСТ: {sheet_name}\n{'='*50}\n")
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            extracted_text.append(
                                " | ".join(str(c) if c is not None else "" for c in row)
                            )
            else:
                import xlrd

                wb = xlrd.open_workbook(file_path)
                extracted_text = []
                for sheet_idx in range(wb.nsheets):
                    sheet = wb.sheet_by_index(sheet_idx)
                    extracted_text.append(f"\n{'='*50}\nЛИСТ: {sheet.name}\n{'='*50}\n")
                    for row_idx in range(sheet.nrows):
                        row = sheet.row_values(row_idx)
                        if any(cell for cell in row):
                            extracted_text.append(
                                " | ".join(str(c) if c else "" for c in row)
                            )

            return "\n".join(extracted_text).strip()

        except ImportError as e:
            return f"Ошибка: для обработки Excel требуется openpyxl/xlrd: {e}"
        except Exception as e:
            logger.error("Excel extraction error: %s", e, exc_info=True)
            return f"Ошибка при чтении Excel: {e!s}"

    @classmethod
    async def extract_structured_data(cls, file_path: str) -> dict[str, Any]:
        """Извлечь структурированные данные: текст + метаданные + статистика."""
        path = Path(file_path)
        ext = path.suffix.lower()
        text = await cls.extract_text_async(file_path)
        stats = {
            "chars": len(text),
            "words": len(text.split()),
            "lines": text.count("\n"),
        }
        file_stat = path.stat()
        metadata = {
            "filename": path.name,
            "extension": ext,
            "size_bytes": file_stat.st_size,
            "size_mb": round(file_stat.st_size / (1024 * 1024), 2),
        }
        tables = None
        if ext in (".xlsx", ".xls"):
            loop = asyncio.get_event_loop()
            tables = await loop.run_in_executor(
                _executor, cls._extract_tables_sync, file_path, ext
            )
        return {"text": text, "metadata": metadata, "stats": stats, "tables": tables}

    @classmethod
    def _extract_tables_sync(cls, file_path: str, ext: str) -> list[dict[str, Any]]:
        """Синхронное извлечение таблиц из Excel."""
        try:
            tables = []
            if ext == ".xlsx":
                import openpyxl

                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    data = [
                        list(row)
                        for row in sheet.iter_rows(values_only=True)
                        if any(c is not None for c in row)
                    ]
                    if data:
                        tables.append({
                            "sheet_name": sheet_name,
                            "headers": data[0],
                            "data": data[1:],
                            "rows_count": len(data) - 1,
                        })
            else:
                import xlrd

                wb = xlrd.open_workbook(file_path)
                for i in range(wb.nsheets):
                    sheet = wb.sheet_by_index(i)
                    data = [
                        sheet.row_values(r)
                        for r in range(sheet.nrows)
                        if any(sheet.row_values(r))
                    ]
                    if data:
                        tables.append({
                            "sheet_name": sheet.name,
                            "headers": data[0],
                            "data": data[1:],
                            "rows_count": len(data) - 1,
                        })
            return tables
        except Exception as e:
            logger.error("Table extraction error: %s", e, exc_info=True)
            return []